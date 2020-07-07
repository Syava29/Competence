from collections import Counter
from openpyxl import load_workbook
import re 
from tkinter import * 
from tkinter import scrolledtext  
import numpy as np


""" Название файлов """
filename1 = "Компетенции.xlsx"
filename_books = "spisok_knig.xlsx"

""" Списки """
spisok = []
spisok_book = []
spisok_description = []
spisok_zuv = []
spisok_content = []
spisok_nazv = []
spisok_komp = []
description_zuv = []

""" Считываем данные из excel """
wb = load_workbook(filename1)
wb_books = load_workbook(filename_books)
sheet_books = wb_books['Лист1']
sheet = wb['Лист1']

""" Загружаем ЗУВы и разбиваем их на слова, а также удаляем Знать, Уметь, Владеть """
for row in sheet['C1':'C35']:
    zuv = ''
    for cell in row:
        zuv = zuv + str(cell.value)
        your_string = zuv
        removal_list = ['Знать','Уметь','Владеть', ' и ', 'для', 'None', ' в ', ' на ',' по ', 'знать', 'Задача', 'Лабораторная', 'Практикум', 'Краткое', 'cодержание', 'Задачи', 'Вариант','Основные', ' с ', ' к ', ' при ']
        for word in removal_list:
            your_string = your_string.replace(word, '')   
    description_zuv.append(your_string)    
    result_key = re.findall(r'\w+', your_string)
    col_count = Counter(result_key).most_common(5)   
    result_main23 = re.sub(r'\d', '', str(col_count))
    res_res = re.findall(r'\w+', result_main23) 
    spisok_zuv.append(res_res)  #Список слов из ЗУВ для каждой компетенции для сравнния с литературой 

for row in sheet['A1':'A35']:
    string = ''
    for cell in row:
        string = string + str(cell.value)
        your_string = string

    spisok_komp.append(your_string)

""" Парсим содержание книг """
sheet_books.cell(row=2, column=1).value
for row in sheet_books['A1':'HK19']:
    string = ''
    for cell in row:
        string = string + str(cell.value) + '\n\n'
        your_string = string
        removal_list = [' и ',' а ',' к ','None' , 'Глава', ' в ', 'Задача', 'Лабораторная', 'Практикум', 'Краткое', 'содержание', 'Задачи', 'Вариант','главы','упражнения', 'работы','Основные']
        for word in removal_list:
            your_string = your_string.replace(word, '')  
    
    result_main = re.sub(r'\d', '', your_string)
    result_main = re.findall(r'\w+', result_main)
    col_count_books = Counter(result_main)
    res_main = re.sub(r'\d', '', str(col_count_books))
    res_main = re.findall(r'\w+', res_main)
    testtt = Counter(result_main) 
    spisok_content.append(res_main)  # Список слов из содержаний книг

""" Cписок названий """
for row in sheet_books['A1':'A19']:
    string = ''
    for cell in row:
        string = string + str(cell.value)
        your_string = string

    spisok_nazv.append(your_string)

""" проверка совпадений самых популярных слов из ЗУВа с каждым содержанием книги """
def clicked():
    txtscrol.delete(1.0, END)
    txt_description.delete(1.0, END)
    i = 0
    k = 0
    #res = "Привет {}".format(txt.get())  
    #lbl.configure(text=res)  
    while i < spisok_zuv.__len__():
        while k < spisok_content.__len__():
            result111=list(set(spisok_content[k]) & set(spisok_zuv[i]))
            if result111.__len__() > 1:
                #print(spisok_komp[i])
                #print(spisok_content[k])
                #print(spisok_zuv[i])
                #print(spisok_nazv[k])
                if txt.get() == spisok_komp[i]:
                    txtscrol.insert(INSERT, 'Для данной компетенции подходит кинга: "' + str(spisok_nazv[k]) + "\n\n")
                    txt_description.insert(INSERT, str(description_zuv[i]) + "\n\n")
                    #lbl.configure(text='Для данной компетенции подходит кинга: "' + str(spisok_nazv[k]) + '"')
                    #print('Для данной компетенции подходит кинга: "' + str(spisok_nazv[k]) + '"')    
            k = k + 1
        
        k = 0
        i = i + 1

""" Визуализация """  
window = Tk()  
window.title("Подбор литературы по компетенциям")  
window.geometry('1500x500')  
lbl_main = Label(window, text = 'Шаг 1')
lbl_main.grid(column=0, row=0)
lbl_main2 = Label(window, text = 'Шаг 2 - нажмите на кнопку')
lbl_main2.grid(column=3, row=0) 
lbl = Label(window, text="Выберите компетенцию для которой необходимо подобрать литературу", font=("Arial Bold", 10))  
lbl.grid(column=0, row=1)  
txt = Entry(window,width=20)  
txt.grid(column=2, row=1)
lbl_description = Label(window, text = 'Знать, уметь, владеть')
lbl_description.grid(column=3, row=3)
txt_description = scrolledtext.ScrolledText(window, width=100, height=10)
txt_description.grid(column=3, row=4) 
txtscrol = scrolledtext.ScrolledText(window, width=100, height=10)  
txtscrol.grid(column=3, row=2)
btn = Button(window, text="Подобрать литературу", command=clicked) 
btn.grid(column=3, row=1)  
 
window.mainloop()