#import openpyxl as excl
from openpyxl import load_workbook
import numpy as np

filename1 = "Компетенции.xlsx"
filename2 = "load.txt"


word = input()
filename = "load.txt"
ffile = open(filename,'r')




text_file = open(filename2, 'w')

wb = load_workbook(filename1)
sheet = wb.get_sheet_by_name('Лист1')
mtrx = np.zeros([100,3])

sheet.cell(row=1, column=2).value
for row in sheet['A1':'C35']:
    string = ''
    for cell in row:
        string = string + str(cell.value) + '\n\n'
    text_file.write(string)



print('Слово "' + word +'" встречается' if word in ffile.read() else 'Нет такого слова')
print(ffile)

text_file.close()

# Print out values in column 2 
#for i in range(1, 4):
   #  print(i, sheet.cell(row=i, column=2).value)

#print(mtrx)